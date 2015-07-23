from openpyxl import load_workbook

logins = {}

def get_logins(filename, sheetname):
    wk = load_workbook(filename)
    ws = wk.get_sheet_by_name(sheetname)
    rowcount = 2            # assumes 1st row contains column headers
    for r in ws.rows:
        username = ws.cell(row = rowcount, column = 2)
        password = ws.cell(row = rowcount, column = 3)
        print(username.value)
        print (password.value)
        rowcount = rowcount + 1
        logins[username.value] = password.value
        try:                            #picks up on an empty row on the end - this removes it
            del logins[None]
        except KeyError:
            pass
    return logins